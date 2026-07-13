from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font

from app.services import calculs


NOMS_MOIS = [
    "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
    "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre",
]


def exporter_mois_excel(annee: int, mois: int, donnees: dict) -> BytesIO:
    wb = Workbook()
    ws_revenus = wb.active
    ws_revenus.title = "Revenus"
    ws_revenus.append(["Registre Mensuel des Revenus (Transport de Personnes)"])
    ws_revenus.append([])
    ws_revenus.append([
        "Date", "Nombre de courses", "Revenu Brut (avant taxes)",
        "Redevance Gouv. (0,90$ / course)", "TPS Perçue (5%)", "TVQ Perçue (9,975%)",
        "Pourboires", "Total Net Encaissé",
    ])
    for revenu in donnees["revenus"]:
        ws_revenus.append([
            revenu["date"].isoformat(),
            revenu["nombre_courses"],
            float(revenu["revenu_brut"]),
            float(revenu["redevance_gouv"]),
            float(revenu["tps_percue"]),
            float(revenu["tvq_percue"]),
            float(revenu["pourboires"]),
            float(revenu["total_net_encaisse"]),
        ])
    if donnees["revenus"]:
        total = {
            "nombre_courses": sum(r["nombre_courses"] for r in donnees["revenus"]),
            "revenu_brut": sum((r["revenu_brut"] for r in donnees["revenus"]), Decimal("0")),
            "redevance_gouv": sum((r["redevance_gouv"] for r in donnees["revenus"]), Decimal("0")),
            "tps_percue": sum((r["tps_percue"] for r in donnees["revenus"]), Decimal("0")),
            "tvq_percue": sum((r["tvq_percue"] for r in donnees["revenus"]), Decimal("0")),
            "pourboires": sum((r["pourboires"] for r in donnees["revenus"]), Decimal("0")),
            "total_net_encaisse": sum((r["total_net_encaisse"] for r in donnees["revenus"]), Decimal("0")),
        }
        ws_revenus.append([
            "Total", total["nombre_courses"], float(total["revenu_brut"]),
            float(total["redevance_gouv"]), float(total["tps_percue"]), float(total["tvq_percue"]),
            float(total["pourboires"]), float(total["total_net_encaisse"]),
        ])

    ws_depenses = wb.create_sheet("Dépenses")
    ws_depenses.append(["Registre Mensuel des Dépenses"])
    ws_depenses.append([])
    ws_depenses.append([
        "Date", "Description du Fournisseur", "Catégorie",
        "Montant Avant Taxes", "TPS Payée", "TVQ Payée", "Montant Total",
    ])
    for depense in donnees["depenses"]:
        ws_depenses.append([
            depense["date"].isoformat(),
            depense["fournisseur"],
            depense["categorie_nom"],
            float(depense["montant_ht"]),
            float(depense["tps"]),
            float(depense["tvq"]),
            float(depense["montant_total"]),
        ])
    if donnees["depenses"]:
        ws_depenses.append([
            "Total", "", "",
            float(sum((d["montant_ht"] for d in donnees["depenses"]), Decimal("0"))),
            float(sum((d["tps"] for d in donnees["depenses"]), Decimal("0"))),
            float(sum((d["tvq"] for d in donnees["depenses"]), Decimal("0"))),
            float(sum((d["montant_total"] for d in donnees["depenses"]), Decimal("0"))),
        ])

    ws_km = wb.create_sheet("Registre Kilométrage")
    ws_km.append(["Registre de Kilométrage"])
    ws_km.append([])
    ws_km.append([
        "Date", "Odomètre Début", "Odomètre Fin",
        "Kilomètres Totaux", "Kilomètres Professionnels", "Taux d'utilisation pro.",
    ])
    for entree in donnees["kilometrage"]["entrees"]:
        ws_km.append([
            entree["date"].isoformat(),
            float(entree["odometre_debut"]),
            float(entree["odometre_fin"]),
            float(entree["km_totaux"]),
            float(entree["km_professionnels"]),
            float(entree["taux_pro"]),
        ])
    totaux = donnees["kilometrage"]["totaux"]
    ws_km.append([
        "Total", "", "",
        float(totaux["km_totaux_mois"]),
        float(totaux["km_pro_mois"]),
        float(totaux["taux_pro"]),
    ])

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
